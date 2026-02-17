"""
Excel Writer Utility for N-gram Automation
Generates multi-sheet Excel output with N-gram analysis per campaign.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, List
import os
from datetime import datetime


# Column configurations for output
NGRAM_COLUMNS = [
    ('ngram', 'N-gram', 20),
    ('impressions', 'Impression', 12),
    ('clicks', 'Click', 10),
    ('spend', 'Spend', 12),
    ('orders', 'Order 14d', 12),
    ('sales', 'Sales 14d', 12),
    ('ctr', 'CTR', 10),
    ('cvr', 'CVR', 10),
    ('cpc', 'CPC', 10),
    ('acos', 'ACOS', 10),
    ('suggestion', 'NE/NP', 8)
]

SEARCH_TERM_COLUMNS = [
    ('search_term', 'Search Term', 35),
    ('impressions', 'Impression', 12),
    ('clicks', 'Click', 10),
    ('spend', 'Spend', 12),
    ('orders', 'Order 14d', 12),
    ('sales', 'Sales 14d', 12),
    ('suggestion', 'NE/NP', 8)
]

# Colors
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONO_HEADER_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
BI_HEADER_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
TRI_HEADER_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
SEARCH_HEADER_FILL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Column width for each N-gram section (number of columns per section + 1 gap)
SECTION_WIDTH = len(NGRAM_COLUMNS) + 1  # 12 columns per section (11 + 1 gap)


def sanitize_sheet_name(name: str, max_length: int = 31) -> str:
    """
    Sanitize sheet name for Excel (max 31 chars, no special chars).
    """
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, ' ')
    
    if len(name) > max_length:
        name = name[:max_length-3] + '...'
    
    return name.strip()


def format_percentage(value) -> str:
    """Format a value as percentage."""
    if pd.isna(value) or value is None:
        return ""
    return f"{value:.2f}%"


def format_currency(value) -> str:
    """Format a value as currency."""
    if pd.isna(value) or value is None:
        return "$0.00"
    return f"${value:.2f}"


def write_ngram_section_horizontal(ws, df: pd.DataFrame, start_row: int, start_col: int,
                                    section_name: str, header_fill: PatternFill, 
                                    columns: list, ref_col: str) -> int:
    """
    Write an N-gram section to a worksheet at a specific column position (horizontal layout).
    
    Args:
        ws: Worksheet object
        df: DataFrame with N-gram data
        start_row: Starting row number
        start_col: Starting column number
        section_name: Name of the section (e.g., "Monogram")
        header_fill: Fill color for headers
        columns: Column configuration list
        ref_col: Column letter for reference keywords list (for NE/NP formula)
        
    Returns:
        Number of rows written (for tracking max height)
    """
    current_row = start_row
    
    # Write section header
    ws.cell(row=current_row, column=start_col, value=section_name)
    ws.cell(row=current_row, column=start_col).font = Font(bold=True, size=12)
    current_row += 2
    
    # Write column headers
    for col_idx, (col_key, col_name, col_width) in enumerate(columns):
        actual_col = start_col + col_idx
        cell = ws.cell(row=current_row, column=actual_col, value=col_name)
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(actual_col)].width = col_width
    
    current_row += 1
    
    # Find the index of columns for formula
    ngram_col_idx = None
    for idx, (col_key, col_name, col_width) in enumerate(columns):
        if col_key == 'ngram' or col_key == 'search_term':
            ngram_col_idx = idx
    
    # Write data rows
    rows_written = 0
    if not df.empty:
        for row_idx, row in df.iterrows():
            for col_idx, (col_key, col_name, col_width) in enumerate(columns):
                actual_col = start_col + col_idx
                value = row.get(col_key, '')
                
                # Format specific columns
                if col_key in ['ctr', 'cvr', 'acos']:
                    value = format_percentage(value) if pd.notna(value) else ''
                elif col_key in ['spend', 'sales', 'cpc']:
                    value = format_currency(value) if pd.notna(value) else ''
                elif col_key in ['impressions', 'clicks', 'orders']:
                    value = int(value) if pd.notna(value) else 0
                elif col_key == 'suggestion':
                    # Use Excel formula for NE/NP based on reference column
                    ngram_cell = get_column_letter(start_col + ngram_col_idx) + str(current_row)
                    # Formula: Check if keyword exists in reference column, mark as "NP"
                    formula = f'=IF(ISNUMBER(MATCH({ngram_cell},{ref_col}:{ref_col},0)),"NP","")'
                    cell = ws.cell(row=current_row, column=actual_col, value=formula)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal='center')
                    continue
                
                cell = ws.cell(row=current_row, column=actual_col, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center' if col_key != 'ngram' and col_key != 'search_term' else 'left')
            
            current_row += 1
            rows_written += 1
    
    return rows_written + 3  # Include header rows


def create_campaign_sheet(wb: Workbook, campaign_name: str, ngrams: Dict[str, pd.DataFrame]) -> None:
    """
    Create a worksheet for a single campaign with N-gram sections side by side (horizontal layout).
    
    Args:
        wb: Workbook object
        campaign_name: Name of the campaign
        ngrams: Dictionary with monograms, bigrams, trigrams, search_terms DataFrames
    """
    sheet_name = sanitize_sheet_name(campaign_name)
    ws = wb.create_sheet(title=sheet_name)
    
    current_row = 1
    
    # Write campaign header
    ws.cell(row=current_row, column=1, value=f"Campaign: {campaign_name}")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=36)
    current_row += 1
    
    # Write back to summary link with hyperlink
    cell = ws.cell(row=current_row, column=1, value="← Back to Summary")
    cell.font = Font(color="0563C1", underline="single")
    cell.hyperlink = "#'Summary'!A1"
    current_row += 2
    
    # Define column positions for horizontal layout
    mono_start_col = 1
    bi_start_col = mono_start_col + SECTION_WIDTH  # After Monogram + gap
    tri_start_col = bi_start_col + SECTION_WIDTH   # After Bigram + gap
    search_start_col = tri_start_col + SECTION_WIDTH  # After Trigram + gap
    
    # Reference columns (far right) - Monogram, Bigram, Trigram
    mono_ref_col_num = search_start_col + len(SEARCH_TERM_COLUMNS) + 2
    bi_ref_col_num = mono_ref_col_num + 1
    tri_ref_col_num = bi_ref_col_num + 1
    
    mono_ref_col = get_column_letter(mono_ref_col_num)
    bi_ref_col = get_column_letter(bi_ref_col_num)
    tri_ref_col = get_column_letter(tri_ref_col_num)
    
    ngram_start_row = current_row
    
    # Get DataFrames
    mono_df = ngrams.get('monograms', pd.DataFrame())
    bi_df = ngrams.get('bigrams', pd.DataFrame())
    tri_df = ngrams.get('trigrams', pd.DataFrame())
    st_df = ngrams.get('search_terms', pd.DataFrame())
    
    # Write Monogram section - references Monogram ref column
    mono_rows = write_ngram_section_horizontal(ws, mono_df, ngram_start_row, mono_start_col,
                                                "Monogram", MONO_HEADER_FILL, NGRAM_COLUMNS,
                                                mono_ref_col)
    
    # Write Bigram section - references Bigram ref column
    bi_rows = write_ngram_section_horizontal(ws, bi_df, ngram_start_row, bi_start_col,
                                              "Bigram", BI_HEADER_FILL, NGRAM_COLUMNS,
                                              bi_ref_col)
    
    # Write Trigram section - references Trigram ref column
    tri_rows = write_ngram_section_horizontal(ws, tri_df, ngram_start_row, tri_start_col,
                                               "Trigram", TRI_HEADER_FILL, NGRAM_COLUMNS,
                                               tri_ref_col)
    
    # Write Search Term section - references all three columns (check mono, bi, tri)
    # For search terms, we'll check all three reference columns
    search_rows = 0
    if not st_df.empty:
        # Custom handling for search terms - check all three ref columns
        search_rows = write_search_term_section_with_refs(ws, st_df, ngram_start_row, search_start_col,
                                                           SEARCH_HEADER_FILL, SEARCH_TERM_COLUMNS,
                                                           mono_ref_col, bi_ref_col, tri_ref_col)
    
    # Create reference columns at the end (Monogram, Bigram, Trigram)
    # Monogram reference column (green)
    ws.cell(row=ngram_start_row, column=mono_ref_col_num, value="Monogram")
    ws.cell(row=ngram_start_row, column=mono_ref_col_num).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ngram_start_row, column=mono_ref_col_num).fill = MONO_HEADER_FILL
    ws.column_dimensions[mono_ref_col].width = 25
    
    ws.cell(row=ngram_start_row + 1, column=mono_ref_col_num, value="(Add keywords to negate)")
    ws.cell(row=ngram_start_row + 1, column=mono_ref_col_num).font = Font(italic=True, size=9, color="666666")
    
    # Bigram reference column (orange)
    ws.cell(row=ngram_start_row, column=bi_ref_col_num, value="Bigram")
    ws.cell(row=ngram_start_row, column=bi_ref_col_num).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ngram_start_row, column=bi_ref_col_num).fill = BI_HEADER_FILL
    ws.column_dimensions[bi_ref_col].width = 25
    
    ws.cell(row=ngram_start_row + 1, column=bi_ref_col_num, value="(Add keywords to negate)")
    ws.cell(row=ngram_start_row + 1, column=bi_ref_col_num).font = Font(italic=True, size=9, color="666666")
    
    # Trigram reference column (blue)
    ws.cell(row=ngram_start_row, column=tri_ref_col_num, value="Trigram")
    ws.cell(row=ngram_start_row, column=tri_ref_col_num).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ngram_start_row, column=tri_ref_col_num).fill = TRI_HEADER_FILL
    ws.column_dimensions[tri_ref_col].width = 25
    
    ws.cell(row=ngram_start_row + 1, column=tri_ref_col_num, value="(Add keywords to negate)")
    ws.cell(row=ngram_start_row + 1, column=tri_ref_col_num).font = Font(italic=True, size=9, color="666666")


def write_search_term_section_with_refs(ws, df: pd.DataFrame, start_row: int, start_col: int,
                                         header_fill: PatternFill, columns: list,
                                         mono_ref_col: str, bi_ref_col: str, tri_ref_col: str) -> int:
    """
    Write search terms section.
    """
    current_row = start_row
    
    # Write section header
    ws.cell(row=current_row, column=start_col, value="Search Term")
    ws.cell(row=current_row, column=start_col).font = Font(bold=True, size=12)
    current_row += 2
    
    # Write column headers
    for col_idx, (col_key, col_name, col_width) in enumerate(columns):
        actual_col = start_col + col_idx
        cell = ws.cell(row=current_row, column=actual_col, value=col_name)
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(actual_col)].width = col_width
    
    current_row += 1
    
    # Write data rows
    rows_written = 0
    if not df.empty:
        for row_idx, row in df.iterrows():
            for col_idx, (col_key, col_name, col_width) in enumerate(columns):
                actual_col = start_col + col_idx
                value = row.get(col_key, '')
                
                if col_key in ['ctr', 'cvr', 'acos']:
                    value = format_percentage(value) if pd.notna(value) else ''
                elif col_key in ['spend', 'sales', 'cpc']:
                    value = format_currency(value) if pd.notna(value) else ''
                elif col_key in ['impressions', 'clicks', 'orders']:
                    value = int(value) if pd.notna(value) else 0
                elif col_key == 'suggestion':
                    # Leave NE/NP column empty (no formula)
                    cell = ws.cell(row=current_row, column=actual_col, value='')
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal='center')
                    continue
                
                cell = ws.cell(row=current_row, column=actual_col, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center' if col_key != 'search_term' else 'left')
            
            current_row += 1
            rows_written += 1
    
    return rows_written + 3


def create_summary_sheet(wb: Workbook, campaigns: Dict[str, dict]) -> None:
    """
    Create the summary sheet with list of all campaigns and their metrics.
    """
    ws = wb.active
    ws.title = "Summary"
    
    current_row = 1
    
    # Write title
    ws.cell(row=current_row, column=1, value="N-Gram Analysis Summary")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    current_row += 1
    
    # Write generation date
    ws.cell(row=current_row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=current_row, column=1).font = Font(italic=True)
    current_row += 2
    
    # Write summary headers
    summary_columns = [
        ('Campaign', 40),
        ('Search Terms', 15),
        ('Monograms', 12),
        ('Bigrams', 12),
        ('Trigrams', 12),
        ('Total Spend', 15),
        ('Total Sales', 15)
    ]
    
    for col_idx, (col_name, col_width) in enumerate(summary_columns, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    current_row += 1
    
    # Write campaign rows with hyperlinks to their sheets
    for campaign_name, summary in campaigns.items():
        sheet_name = sanitize_sheet_name(campaign_name)
        
        # Campaign name cell with hyperlink to the campaign sheet
        cell = ws.cell(row=current_row, column=1, value=campaign_name)
        cell.border = THIN_BORDER
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = Font(color="0563C1", underline="single")
        
        ws.cell(row=current_row, column=2, value=summary.get('search_term_count', 0)).border = THIN_BORDER
        ws.cell(row=current_row, column=3, value=summary.get('monogram_count', 0)).border = THIN_BORDER
        ws.cell(row=current_row, column=4, value=summary.get('bigram_count', 0)).border = THIN_BORDER
        ws.cell(row=current_row, column=5, value=summary.get('trigram_count', 0)).border = THIN_BORDER
        ws.cell(row=current_row, column=6, value=format_currency(summary.get('total_spend', 0))).border = THIN_BORDER
        ws.cell(row=current_row, column=7, value=format_currency(summary.get('total_sales', 0))).border = THIN_BORDER
        
        # Center align numeric columns
        for col_idx in range(2, 8):
            ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center')
        
        current_row += 1
    
    # Add totals row
    current_row += 1
    ws.cell(row=current_row, column=1, value="TOTAL").font = Font(bold=True)


def create_excel_output(processed_data: Dict[str, Dict], output_path: str) -> str:
    """
    Create the final Excel output file with all campaigns.
    """
    wb = Workbook()
    
    campaign_summaries = {}
    
    for campaign_name, data in processed_data.items():
        ngrams = data.get('ngrams', {})
        summary = data.get('summary', {})
        
        create_campaign_sheet(wb, campaign_name, ngrams)
        campaign_summaries[campaign_name] = summary
    
    create_summary_sheet(wb, campaign_summaries)
    wb.save(output_path)
    
    return output_path


def generate_output_filename(prefix: str = "NGram_Analysis") -> str:
    """Generate a filename with timestamp."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"{prefix}_{timestamp}.xlsx"


def create_asin_report(df_asins: pd.DataFrame, output_path: str) -> str:
    """
    Create an Excel report containing only ASIN data with their stats.
    
    Args:
        df_asins: DataFrame containing only ASIN rows
        output_path: Path to save the Excel file
        
    Returns:
        Path to the created file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ASIN Report"
    
    current_row = 1
    
    # Write title
    ws.cell(row=current_row, column=1, value="ASIN Targeting Report")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    current_row += 1
    
    # Write generation date
    ws.cell(row=current_row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=current_row, column=1).font = Font(italic=True)
    current_row += 2
    
    # Define columns for ASIN report
    asin_columns = [
        ('search_term', 'ASIN', 15),
        ('campaign', 'Campaign', 40),
        ('ad_group', 'Ad Group', 30),
        ('impressions', 'Impressions', 12),
        ('clicks', 'Clicks', 10),
        ('spend', 'Spend', 12),
        ('sales', 'Sales', 12),
        ('orders', 'Orders', 10),
        ('acos', 'ACOS', 10),
    ]
    
    # Header fill for ASIN report
    ASIN_HEADER_FILL = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
    
    # Write headers
    for col_idx, (col_key, col_name, col_width) in enumerate(asin_columns, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.fill = ASIN_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    current_row += 1
    
    # Write data rows (all rows without grouping/deduplication)
    if not df_asins.empty:
        # Calculate ACOS for each row
        df_asins = df_asins.copy()
        df_asins['acos'] = df_asins.apply(
            lambda row: (row['spend'] / row['sales'] * 100) if row['sales'] > 0 else 0, 
            axis=1
        )
        
        # Sort by spend descending
        df_asins = df_asins.sort_values('spend', ascending=False)
        
        for _, row in df_asins.iterrows():
            for col_idx, (col_key, col_name, col_width) in enumerate(asin_columns, 1):
                value = row.get(col_key, '')
                
                # Format specific columns
                if col_key == 'acos':
                    value = format_percentage(value) if pd.notna(value) else ''
                elif col_key in ['spend', 'sales']:
                    value = format_currency(value) if pd.notna(value) else ''
                elif col_key in ['impressions', 'clicks', 'orders']:
                    value = int(value) if pd.notna(value) else 0
                
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center' if col_key not in ['search_term', 'campaign', 'ad_group'] else 'left')
            
            current_row += 1
    
    # Add summary row
    current_row += 1
    ws.cell(row=current_row, column=1, value="TOTAL").font = Font(bold=True)
    
    if not df_asins.empty:
        ws.cell(row=current_row, column=4, value=int(df_asins['impressions'].sum())).font = Font(bold=True)
        ws.cell(row=current_row, column=5, value=int(df_asins['clicks'].sum())).font = Font(bold=True)
        ws.cell(row=current_row, column=6, value=format_currency(df_asins['spend'].sum())).font = Font(bold=True)
        ws.cell(row=current_row, column=7, value=format_currency(df_asins['sales'].sum())).font = Font(bold=True)
        if 'orders' in df_asins.columns:
            ws.cell(row=current_row, column=8, value=int(df_asins['orders'].sum())).font = Font(bold=True)
        
        total_spend = df_asins['spend'].sum()
        total_sales = df_asins['sales'].sum()
        total_acos = (total_spend / total_sales * 100) if total_sales > 0 else 0
        ws.cell(row=current_row, column=9, value=format_percentage(total_acos)).font = Font(bold=True)
    
    wb.save(output_path)
    return output_path
